import openpyxl
import os
from datetime import datetime
import xml.etree.ElementTree as ET
import keyboard


# Путь к файлу xlsx
FilePaths = 'E:\\1\\'
FileXmlName = ''
# Товары
product = []
# Итоговая сумма документа
totalAmount = 0
# Итоговая сумма с расчетом НДС 20%
totalAmountNds = 0
# Сумма налога от Итоговой суммы
totalAmountTax = 0
# Счетчик
count = 1


def getFiles(filePaths):
    global FileXmlName 
    filesXlsx = None
    fileXml = None
    files = os.listdir(filePaths)

# Получение всех файлов в указанной директории, исключая файлы, содержащие 'NEW_' в названии
    files = [file for file in os.listdir(filePaths) if os.path.isfile(os.path.join(filePaths, file)) and 'NEW_' not in file]

# Сортировка файлов по дате изменения
    files.sort(key=lambda file: os.path.getmtime(os.path.join(filePaths, file)), reverse=True)

# Получение двух самых новых файлов
    last_two_files = files[:2]

    for file in last_two_files:
        if file.endswith('.xlsx'):
            if file.startswith('~$'):
                print(f'Закройте пожалуйста файл {file}... и нажмите ENTER для продолжения')
                keyboard.read_event()
                return getFiles(FilePaths)
            filesXlsx = os.path.join(filePaths, file)
        elif file.endswith('.xml'):
            FileXmlName = file
            fileXml = os.path.join(filePaths, file)
    retrievingDataFromXlsx(filesXlsx)
    createXml(fileXml)


# В данной функции получаем данные из файла и заполняем словари
def retrievingDataFromXlsx(filesXlsx):
    global totalAmount
    wookbook = openpyxl.load_workbook(filesXlsx, data_only=True)
    worksheet = wookbook.active
    for i in range(1, worksheet.max_row):
        if worksheet['B' + str(i)].value == 'Наименование оборудования и Работ':
            for j in range(i+1, worksheet.max_row):
                if worksheet['B' + str(j)].value == None:
                    totalAmount = worksheet['K' + str(j)].value
                    break
                product.append({'product': worksheet['B' + str(j)].value,
                                'model': worksheet['F' + str(j)].value,
                                'number': worksheet['H' + str(j)].value,
                                'unit': worksheet['I' + str(j)].value,
                                'price': worksheet['J' + str(j)].value,
                                'sum': worksheet['K' + str(j)].value,
                                'sumNds': worksheet['K' + str(j)].value * 1.2,
                                'taxSumNds': worksheet['K' + str(j)].value * 1.2 - worksheet['K' + str(j)].value})


# Собираем новый XML файл из полученных данных
def createXml(fileXml):
    global count
    global totalAmountNds
    global totalAmountTax

    totalAmountNds = totalAmount * 1.2
    totalAmountTax = totalAmountNds - totalAmount

    tree = ET.parse(fileXml)
    root = tree.getroot()

    document_node = root.find('.//Документ')

    svschet_fact_node = document_node.find('.//СвСчФакт')
    if svschet_fact_node is None:
        svschet_fact_node = ET.SubElement(document_node, 'СвСчФакт')

    for tablschfact in document_node.findall('ТаблСчФакт'):
        document_node.remove(tablschfact)

    newTable = ET.Element('ТаблСчФакт')
    index = list(document_node).index(svschet_fact_node) + 1

    for products in product:
        new_svedtov = ET.SubElement(newTable, 'СведТов', {
            'НомСтр': f"{count}",
            'НаимТов': f"{products['product']}" + ' ' + f"{products['model']}" if products['model'] != None else f"{products['product']}",
            'ОКЕИ_Тов': "796",
            'КолТов': f"{products['number']}",
            'ЦенаТов': f"{products['price']:.2f}",
            'СтТовБезНДС': f"{products['sum']:.2f}",
            'НалСт': "20%",
            'СтТовУчНал': f"{products['sumNds']:.2f}"
        })
        akziz = ET.SubElement(new_svedtov, 'Акциз')
        bezakziz = ET.SubElement(akziz, 'БезАкциз')
        bezakziz.text = 'без акциза'

        sumnal = ET.SubElement(new_svedtov, 'СумНал')
        sumnal_value = ET.SubElement(sumnal, 'СумНал')
        sumnal_value.text = f"{products['taxSumNds']:.2f}"

        dopsvedtov = ET.SubElement(new_svedtov, 'ДопСведТов', {
            'НаимЕдИзм': f"{products['unit']}"
        })
        count = count + 1
    
    new_vsogoopl = ET.SubElement(newTable, 'ВсегоОпл', {
    'СтТовБезНДСВсего': f"{totalAmount:.2f}",
    'СтТовУчНалВсего': f"{totalAmountNds:.2f}"
    })
    sumnalvsego = ET.SubElement(new_vsogoopl, 'СумНалВсего')
    sumnal = ET.SubElement(sumnalvsego, 'СумНал')
    sumnal.text = f'{totalAmountTax:.2f}'
    kolnetto = ET.SubElement(new_vsogoopl, 'КолНеттоВс')
    kolnetto.text = '106.000'

    document_node.insert(index, newTable)


    # Сохраняем изменения в XML-файл
    tree.write(f'{FilePaths}' + 'NEW_' + f'{FileXmlName}', encoding='utf-8', xml_declaration=True)



getFiles(FilePaths)

print('XML файл успешно создан. Для закрытия программы нажмите любую клавишу.')

keyboard.read_event()

print("программа завершена")
